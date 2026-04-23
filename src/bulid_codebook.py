"""
마스터 코드북 생성 & 데이터 표준화 스크립트

출력:
  data/codebook/master_codebook.xlsx
    - [변수사전] 표준컬럼명 | 한글설명 | 타입 | 비고 | 연도별 원본컬럼명
    - [값레이블] 표준컬럼명 | 한글설명 | 코드값 | 레이블
  data/processed/master_standardized.csv  (숫자전용)

실행: python src/build_codebook.py
"""

import pandas as pd
import numpy as np
import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

MAPPING_FILE = Path("data/codebook/variable_mapping_table_final.xlsx")
MASTER_CSV   = Path("data/processed/master.csv")
OUT_CODEBOOK = Path("data/codebook/master_codebook.xlsx")
OUT_CSV      = Path("data/processed/master_standardized.csv")

HEADER_ROW = {2016:0,2017:0,2018:1,2019:1,2020:1,2021:1,2022:1,2023:2,2024:1,2025:4}
YEARS      = list(HEADER_ROW.keys())

# ── 종목 코드 (2025 가계지출 코드북 기준) ──────────────────────────────────────
SPORT_CODES = {
    101:'골프', 102:'스크린골프', 103:'게이트볼', 104:'그라운드골프', 105:'파크골프',
    106:'농구', 107:'당구/포켓볼', 108:'라켓볼', 109:'미식축구, 럭비',
    110:'배구', 111:'배드민턴', 112:'볼링', 113:'소프트볼', 114:'수구',
    115:'스쿼시', 116:'야구', 117:'정구', 118:'족구', 119:'축구', 120:'풋살',
    121:'탁구', 122:'테니스', 123:'하키(인라인, 필드)', 124:'핸드볼',
    201:'걷기', 202:'조깅', 203:'육상 마라톤', 204:'댄스스포츠',
    205:'벨리댄스', 206:'재즈댄스', 207:'보디빌딩(헬스)', 208:'생활체조',
    209:'수영', 210:'아쿠아로빅/수중발레', 211:'에어로빅', 212:'요가',
    213:'자전거', 214:'줄넘기', 215:'크로스핏', 216:'필라테스', 217:'훌라후프',
    301:'복싱', 302:'격투기(킥복싱, 이종격투기)', 303:'검도', 304:'유도',
    305:'레슬링', 306:'태권도', 307:'합기도',
    308:'무도(유도, 검도, 태권도, 합기도 제외)', 309:'펜싱',
    401:'국궁/석궁/양궁', 402:'씨름',
    501:'래프팅', 502:'빙상', 503:'사격', 504:'서바이벌',
    505:'수상스키', 506:'스키/보드', 507:'스킨스쿠버', 508:'승마',
    509:'암벽등반(클라이밍)', 510:'요트', 511:'윈드서핑',
    512:'인라인스케이트', 513:'철인3종', 514:'카누',
    515:'항공레저(스카이다이빙, 패러글라이딩 등)',
    601:'낚시', 602:'등산', 603:'캠핑', 604:'기타',
    701:'홈트레이닝',
}

# ── 변수 설명 ──────────────────────────────────────────────────────────────────
VAR_DESC = {
    "SURVEY_YEAR": ("조사연도",       "숫자", "2016~2025"),
    "ID":          ("응답자 일련번호", "숫자", ""),
    "CODE1":       ("조사구번호",      "문자", ""),
    "CODE2":       ("가구번호",        "숫자", ""),
    "CODE3":       ("시도",            "숫자", "지역코드 → 값레이블 시트"),
    "CODE4":       ("시군구",          "문자", ""),
    "CODE5":       ("읍면동",          "문자", ""),
    "CODE6":       ("읍면동 구분",     "숫자", "1=동부, 2=읍면부"),
    "APT":         ("주거유형",        "숫자", "1=아파트, 2=단독/연립/다세대, 3=기타"),
    "SEX":         ("성별",            "숫자", "1=남성, 2=여성"),
    "AGE":         ("만 나이",         "숫자", "만 나이(세)"),
    "YEAR":        ("출생연도",        "숫자", "YYYY"),
    "MON":         ("출생월",          "숫자", "1~12"),
    "Q1":          ("본인 건강상태 인식",        "숫자", "1~5척도"),
    "Q2":          ("건강상태 유지 중요 요인",    "숫자", "값레이블 참조"),
    "Q3":          ("본인 체력상태 인식",        "숫자", "1~5척도"),
    "Q4":          ("체력상태 유지 중요 요인",    "숫자", "값레이블 참조"),
    "Q5_1":        ("건강체력 유지 방법 - 규칙적 체육활동",     "숫자", "1~5척도"),
    "Q5_2":        ("건강체력 유지 방법 - 충분한 휴식 및 수면", "숫자", "1~5척도"),
    "Q5_3":        ("건강체력 유지 방법 - 규칙적 식사 및 영양보충", "숫자", "1~5척도"),
    "Q5_4":        ("건강체력 유지 방법 - 금주 및 금연",       "숫자", "1~5척도"),
    "Q5_5":        ("건강체력 유지 방법 - 기타",               "숫자", "1~5척도"),
    "Q6":          ("생활권 주변 체육시설 인지 여부",           "숫자", "1=안다, 2=모른다"),
    "Q6_1":        ("최근 1년간 참여 체육활동 종목 1",         "숫자", "종목코드 → 값레이블"),
    "Q6_2":        ("최근 1년간 참여 체육활동 종목 2",         "숫자", "종목코드 → 값레이블"),
    "Q7":          ("규칙적 체육활동 참여 주기 (2025기준)",    "숫자", "값레이블 참조"),
    "Q16":         ("최근 1년간 규칙적 체육활동 참여 빈도",     "숫자", "1~9 → 값레이블"),
    "Q8_1_1":      ("주로 참여 체육활동 1순위 종목",  "숫자", "종목코드 → 값레이블"),
    "Q8_1_2":      ("주로 참여 체육활동 1순위 참여빈도",       "숫자", "값레이블 참조"),
    "Q8_1_3":      ("주로 참여 체육활동 1순위 참여요일",       "숫자", "1=평일, 2=휴일, 3=평일+휴일"),
    "Q8_1_4":      ("주로 참여 체육활동 1순위 주 시간대",      "숫자", "값레이블 참조"),
    "Q8_1_6":      ("주로 참여 체육활동 1순위 최근 1년간 참여기간", "숫자", "값레이블 참조"),
    "Q8_1_7":      ("주로 참여 체육활동 1순위 참여강도",       "숫자", "1=저, 2=중, 3=고"),
    "Q8_1_8":      ("주로 참여 체육활동 1순위 주 이용시설",    "숫자", "값레이블 참조"),
    "Q8_1_9":      ("주로 참여 체육활동 1순위 이동수단",       "숫자", "값레이블 참조"),
    "Q8_2_1":      ("주로 참여 체육활동 2순위 종목",  "숫자", "종목코드 → 값레이블"),
    "Q8_3_1":      ("주로 참여 체육활동 3순위 종목",  "숫자", "종목코드 → 값레이블"),
    "DQ1_1":       ("최종학력",                     "숫자", "값레이블 참조"),
    "DQ1_2":       ("이수여부",                     "숫자", "1=졸업, 2=재학, 3=수료, 4=휴학, 5=중퇴"),
    "DQ2_1":       ("혼인상태",                     "숫자", "값레이블 참조"),
    "DQ2_2":       ("가구원수",                     "숫자", "명"),
    "DQ2_3":       ("자녀현황",                     "숫자", "명"),
    "DQ3":         ("월평균 가구소득",               "숫자", "2016=코드(값레이블참조), 2018+=만원 실수치"),
    "DQ4":         ("직업 유무",                    "숫자", "1=있음, 2=없음"),
    "DQ4_1":       ("무직자 상태",                  "숫자", "값레이블 참조"),
    "DQ4A":        ("직업명",                       "문자", ""),
    "WT":          ("가중치",                       "숫자", ""),
}

# ── 값 레이블 ──────────────────────────────────────────────────────────────────
VALUE_LABELS = {
    "SEX": {1:"남성", 2:"여성"},
    "CODE3": {
        11:"서울특별시", 21:"부산광역시", 22:"대구광역시",
        23:"인천광역시", 24:"광주광역시", 25:"대전광역시",
        26:"울산광역시", 29:"세종특별자치시",
        31:"경기도",     32:"강원도",      33:"충청북도",
        34:"충청남도",   35:"전라북도",    36:"전라남도",
        37:"경상북도",   38:"경상남도",    39:"제주특별자치도",
    },
    "CODE6": {1:"동부(도시)", 2:"읍면부(농촌)"},
    "APT":   {1:"아파트", 2:"단독/연립/다세대", 3:"기타"},
    "Q1": {
        1:"전혀 건강하지 않다", 2:"건강하지 않다",
        3:"보통",              4:"건강한 편이다",  5:"매우 건강하다",
    },
    "Q3": {
        1:"전혀 체력이 없다", 2:"체력이 없다",
        3:"보통",            4:"체력이 있다",   5:"매우 체력이 있다",
    },
    "Q5_1": {1:"전혀 안함",2:"안함",3:"보통",4:"잘 수행",5:"매우 잘 수행"},
    "Q5_2": {1:"전혀 안함",2:"안함",3:"보통",4:"잘 수행",5:"매우 잘 수행"},
    "Q5_3": {1:"전혀 안함",2:"안함",3:"보통",4:"잘 수행",5:"매우 잘 수행"},
    "Q5_4": {1:"전혀 안함",2:"안함",3:"보통",4:"잘 수행",5:"매우 잘 수행"},
    "Q2": {
        1:"규칙적인 체육활동", 2:"충분한 휴식 및 수면",
        3:"규칙적인 식사 및 영양보충", 4:"금주 및 금연",
        5:"기타", 6:"긍정적인 사고방식", 7:"정기검진",
    },
    "Q4": {
        1:"규칙적인 체육활동", 2:"충분한 휴식 및 수면",
        3:"규칙적인 식사 및 영양보충", 4:"금주 및 금연",
        5:"기타", 6:"긍정적인 사고방식",
    },
    # 종목 코드 (Q6_1, Q6_2, Q8_1_1, Q8_2_1, Q8_3_1 공통)
    "종목코드": SPORT_CODES,
    "Q6_1":    SPORT_CODES,
    "Q6_2":    SPORT_CODES,
    "Q8_1_1":  SPORT_CODES,
    "Q8_2_1":  SPORT_CODES,
    "Q8_3_1":  SPORT_CODES,
    "Q7": {
        1:"거의 매일(주5회 이상)", 2:"주 3~4회", 3:"주 1~2회",
        4:"월 1~3회", 5:"분기 1~2회", 6:"반기 1회",
        7:"연 1회 미만", 8:"참여하지 않음", 9:"기타",
    },
    "Q16": {
        1:"전혀하지 않는다", 2:"한달에 3번 이하",
        3:"주1번", 4:"주2번", 5:"주3번",
        6:"주4번", 7:"주5번", 8:"주6번", 9:"매일",
    },
    "Q8_1_2": {
        1:"월3번 이하", 2:"주1번", 3:"주2번", 4:"주3번",
        5:"주4번", 6:"주5번", 7:"주6번", 8:"매일",
    },
    "Q8_1_3": {1:"평일", 2:"휴일(주말)", 3:"평일+휴일"},
    "Q8_1_4": {
        1:"아침/새벽(6~8시)", 2:"오전(8~12시)", 3:"점심(12~14시)",
        4:"오후(14~18시)",    5:"저녁(18~22시)", 6:"일정하지 않음",
    },
    "Q8_1_6": {
        1:"1개월 미만", 2:"1~3개월 미만", 3:"3~6개월 미만",
        4:"6개월~1년 미만", 5:"1~2년 미만", 6:"2~3년 미만",
        7:"3~5년 미만", 8:"5~10년 미만", 9:"10년 이상",
    },
    "Q8_1_7": {1:"저강도", 2:"중강도", 3:"고강도"},
    "Q8_1_8": {
        1:"공공 체육시설", 2:"민간 체육시설",
        3:"학교/직장 체육시설", 4:"기타 부대시설",
        5:"자가시설", 6:"자연환경", 8:"없다",
    },
    "Q8_1_9": {
        1:"도보", 2:"자전거", 3:"오토바이",
        4:"자가용", 5:"택시", 6:"지하철", 7:"버스", 98:"없음",
    },
    "DQ1_1": {
        1:"무학", 2:"초등학교", 3:"중학교", 4:"고등학교",
        5:"대학(4년제 미만)", 6:"대학교(4년제 이상)",
        7:"대학원 석사과정", 8:"대학원 박사과정",
    },
    "DQ1_2": {1:"졸업", 2:"재학", 3:"수료", 4:"휴학", 5:"중퇴"},
    "DQ2_1": {1:"기혼(유배우)", 2:"미혼", 3:"사별", 4:"이혼", 5:"기타"},
    "DQ4":   {1:"있음", 2:"없음"},
    "DQ4_1": {1:"전업주부", 2:"학생", 3:"무직", 4:"기타"},
}

REGION_TO_CODE = {v: k for k, v in VALUE_LABELS["CODE3"].items()}

# ── 변수 매핑 ──────────────────────────────────────────────────────────────────
def load_var_mapping():
    raw = pd.read_excel(MAPPING_FILE, sheet_name="변수매핑테이블", header=None)
    years_order = [2025,2024,2023,2022,2021,2020,2019,2018,2017,2016]
    data = raw.iloc[2:].reset_index(drop=True)
    std_cols = [str(r).strip() for r in data.iloc[:,0]
                if pd.notna(r) and str(r).strip() not in ("nan","해당없음")]
    std_to_orig = {}
    for _, row in data.iterrows():
        std = str(row.iloc[0]).strip()
        if not std or std in ("nan","해당없음"): continue
        std_to_orig.setdefault(std, {})
        for i, yr in enumerate(years_order):
            orig = str(row.iloc[i*2]).strip()
            if orig and orig not in ("nan","해당없음"):
                std_to_orig[std][yr] = orig
    return std_cols, std_to_orig

# ── 표준화 ─────────────────────────────────────────────────────────────────────
def region_to_num(val):
    if pd.isna(val): return np.nan
    s = str(val).strip()
    try:
        n = int(float(s))
        if n in VALUE_LABELS["CODE3"]: return n
    except: pass
    cleaned = re.sub(r'^\d+\.\s*','',s).strip()
    for name, code in REGION_TO_CODE.items():
        if cleaned == name or (len(cleaned)>=2 and cleaned[:2] == name[:2]):
            return code
    return np.nan

def standardize_and_save():
    df = pd.read_csv(MASTER_CSV, dtype=str)
    # 결측 90%+ 제거
    drop_cols = df.columns[df.isna().mean() >= 0.9].tolist()
    df = df.drop(columns=drop_cols)
    # 파생 라벨 컬럼 제거
    df = df.drop(columns=[c for c in df.columns
                           if c.endswith('_LABEL') or c.endswith('_GROUP')], errors='ignore')
    # CODE3 → 숫자
    if "CODE3" in df.columns:
        df["CODE3"] = df["CODE3"].apply(region_to_num)
    # 전체 숫자 변환
    # 코드북에 숫자코드 없는 텍스트 컬럼 제거
    text_drop = [c for c in ["CODE1","CODE4","CODE5","DQ4A"] if c in df.columns]
    df = df.drop(columns=text_drop)
    if text_drop:
        print(f"  텍스트 컬럼 제거: {text_drop}")

    # 나머지 전체 숫자 변환
    for col in df.columns:
        df[col] = pd.to_numeric(df[col], errors='coerce')
    # 이상값 제거
    if "AGE"  in df.columns: df.loc[~df["AGE"].between(10,100), "AGE"]  = np.nan
    if "YEAR" in df.columns: df.loc[~df["YEAR"].between(1920,2015),"YEAR"] = np.nan
    for c in ["Q16","Q7"]:
        if c in df.columns: df.loc[~df[c].between(1,9),c] = np.nan
    df.to_csv(OUT_CSV, index=False, encoding="utf-8-sig")
    print(f"✓ 표준화 CSV: {OUT_CSV}  ({len(df):,}행 × {len(df.columns)}열)")
    return df

# ── Excel 코드북 ───────────────────────────────────────────────────────────────
def build_excel(df_std, std_cols, std_to_orig):
    wb = Workbook()
    thin  = Side(style='thin',  color='CCCCCC')
    thick = Side(style='medium',color='888888')
    bT = Border(left=thin,right=thin,top=thin,bottom=thin)
    ac = Alignment(horizontal='center', vertical='center', wrap_text=True)
    al = Alignment(horizontal='left',   vertical='center', wrap_text=True)

    H1  = PatternFill('solid', fgColor='1F4E79')
    H2  = PatternFill('solid', fgColor='2E75B6')
    KEY = PatternFill('solid', fgColor='E2EFDA')
    ALT = PatternFill('solid', fgColor='F5F9FF')
    SPT = PatternFill('solid', fgColor='FFF8E1')

    KEY_VARS = {"SEX","AGE","Q1","Q16","Q7","DQ1_1","DQ3","DQ4","CODE3",
                "Q6_1","Q8_1_1","Q8_1_2"}
    SPORT_VARS = {"Q6_1","Q6_2","Q8_1_1","Q8_2_1","Q8_3_1"}

    # ── Sheet1: 변수사전 ───────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = '변수사전'
    hdrs = ['#','표준컬럼명','한글설명','데이터유형','비고'] + [str(y) for y in YEARS]

    ws1.append([])
    for ci, h in enumerate(hdrs, 1):
        cell = ws1.cell(row=1, column=ci, value=h)
        cell.font      = Font(name='Arial', bold=True, size=10, color='FFFFFF')
        cell.fill      = H1 if ci <= 5 else H2
        cell.alignment = ac; cell.border = bT

    active = [c for c in std_cols if c in df_std.columns or c in VAR_DESC]
    for ri, std in enumerate(active, 2):
        desc, dtype, note = VAR_DESC.get(std, ('','',''))
        origs = [std_to_orig.get(std,{}).get(yr,'') for yr in YEARS]
        vals  = [ri-1, std, desc, dtype, note] + origs
        is_key   = std in KEY_VARS
        is_sport = std in SPORT_VARS
        fill = KEY if is_key else (SPT if is_sport else (ALT if ri%2==0 else None))
        for ci, v in enumerate(vals, 1):
            cell = ws1.cell(row=ri, column=ci, value=v or None)
            cell.font      = Font(name='Arial', size=9, bold=(is_key and ci<=3))
            cell.alignment = ac if ci<=4 else al
            cell.border    = bT
            if fill: cell.fill = fill

    for ci, w in enumerate([5,18,38,10,32]+[14]*len(YEARS), 1):
        ws1.column_dimensions[get_column_letter(ci)].width = w
    ws1.row_dimensions[1].height = 22
    ws1.freeze_panes = 'A2'

    # ── Sheet2: 값레이블 ───────────────────────────────────────────────────────
    ws2 = wb.create_sheet('값레이블')
    hdrs2 = ['표준컬럼명','한글설명','코드값','레이블']
    ws2.append([])
    for ci, h in enumerate(hdrs2, 1):
        cell = ws2.cell(row=1, column=ci, value=h)
        cell.font=Font(name='Arial',bold=True,size=10,color='FFFFFF')
        cell.fill=H1; cell.alignment=ac; cell.border=bT

    ri2 = 2
    VAR_ORDER = [
        "SEX","CODE3","CODE6","APT",
        "Q1","Q3","Q5_1","Q5_2","Q5_3","Q5_4",
        "Q2","Q4",
        "Q6_1","Q6_2","Q8_1_1","Q8_2_1","Q8_3_1",  # 종목 (공통)
        "Q7","Q16",
        "Q8_1_2","Q8_1_3","Q8_1_4","Q8_1_6","Q8_1_7","Q8_1_8","Q8_1_9",
        "DQ1_1","DQ1_2","DQ2_1","DQ4","DQ4_1",
    ]
    # 중복 종목코드는 "종목코드" 한 번만 출력 후 참조 표시
    sport_printed = False
    for std in VAR_ORDER:
        if std not in VALUE_LABELS: continue
        labels = VALUE_LABELS[std]
        desc   = VAR_DESC.get(std,('','',''))[0]
        is_key   = std in KEY_VARS
        is_sport = std in SPORT_VARS

        # 종목 코드 변수는 대표 한 번 + 나머지는 "→ 종목코드 참조"
        if is_sport and sport_printed:
            cell = ws2.cell(row=ri2, column=1, value=std)
            ws2.cell(row=ri2, column=2, value=desc)
            ws2.cell(row=ri2, column=3, value="101~701")
            ws2.cell(row=ri2, column=4, value="↑ 위 종목코드(Q6_1) 참조")
            for ci in range(1,5):
                ws2.cell(row=ri2,column=ci).font      = Font(name='Arial',size=9,italic=True)
                ws2.cell(row=ri2,column=ci).border    = bT
                ws2.cell(row=ri2,column=ci).alignment = al
                ws2.cell(row=ri2,column=ci).fill      = SPT
            ri2 += 1
            continue

        for code, label in sorted(labels.items(), key=lambda x: x[0]):
            fill2 = KEY if is_key else (SPT if is_sport else (ALT if ri2%2==0 else None))
            vals2 = [std, desc, code, label]
            for ci, v in enumerate(vals2, 1):
                cell = ws2.cell(row=ri2, column=ci, value=v)
                cell.font      = Font(name='Arial', size=9, bold=(is_key and ci==1 and code==list(labels.keys())[0]))
                cell.alignment = ac if ci in (1,3) else al
                cell.border    = bT
                if fill2: cell.fill = fill2
            ri2 += 1

        if is_sport: sport_printed = True

    for ci, w in enumerate([18,38,10,40], 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w
    ws2.row_dimensions[1].height = 22
    ws2.freeze_panes = 'A2'

    wb.save(OUT_CODEBOOK)
    total_labels = sum(len(v) for k,v in VALUE_LABELS.items() if k != "종목코드")
    print(f"✓ 마스터 코드북: {OUT_CODEBOOK}")
    print(f"  - 변수사전: {len(active)}개 변수")
    print(f"  - 값레이블: {len(VALUE_LABELS)-1}개 변수, {len(SPORT_CODES)}개 종목코드 포함")

# ── MAIN ───────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print("="*60)
    print("  마스터 코드북 생성 & 데이터 표준화")
    print("="*60+"\n")
    std_cols, std_to_orig = load_var_mapping()
    print(f"▶ 변수 매핑: {len(std_cols)}개 표준 컬럼\n")
    print("▶ 데이터 표준화...")
    df_std = standardize_and_save()
    print("\n▶ 마스터 코드북 생성...")
    build_excel(df_std, std_cols, std_to_orig)
    print(f"\n✓ 완료!\n  데이터: {OUT_CSV}\n  코드북: {OUT_CODEBOOK}")